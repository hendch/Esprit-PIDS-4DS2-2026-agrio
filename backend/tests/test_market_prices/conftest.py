"""Shared fixtures for market-prices test suite.

Provides:
- ``data_dir``       — tmp_path populated with 6 minimal synthetic source files
- ``mock_db_session``— an AsyncMock session whose execute() returns safe defaults
- ``async_client``   — httpx.AsyncClient wired to the FastAPI app with the mock
                       session injected via dependency override
"""
from __future__ import annotations

from unittest.mock import AsyncMock, MagicMock

import pytest
from httpx import ASGITransport, AsyncClient
from jose import jwt

from app.main import app
from app.persistence.db import get_async_session
from app.settings import settings


def _make_test_token() -> str:
    """Generate a valid JWT token for the test user using the app's secret."""
    return jwt.encode(
        {"sub": "test-user-001", "farm_id": "test-farm-001"},
        settings.jwt_secret,
        algorithm=settings.jwt_algorithm,
    )

# ---------------------------------------------------------------------------
# Synthetic source-file fixture
# ---------------------------------------------------------------------------

_REGIONAL_FILES = [
    "evolution-du-prix-moyen-des-brebis-suitees-par-grande-region-en-dinar.xlsx",
    "evolution-du-prix-moyen-des-genisses-pleines-par-grande-region-en-dinar.xlsx",
    "evolution-du-prix-moyen-des-vaches-suitees-par-grande-region-en-dinar.xlsx",
]

_MEAT_FILE = (
    "evolution-du-prix-moyen-de-vente-des-viandes-rouge-en-dinar-par-kilogramme-vif.xlsx"
)

_XLS_HTML_FILES = [
    "Évolution des prix des bovins suivis par tête, toutes races confondues.xls",
    "Évolution des prix des vaches gestantes par tête.xls",
]


@pytest.fixture
def data_dir(tmp_path):
    """Create a temp directory with 6 minimal synthetic source files (3 rows each).

    Regional XLSX files have columns: annee | mois | nord | sahel | centre-et-sud
    Meat XLSX has columns:            annee | mois | taurillon-maigre | ... | caprin
    HTML-disguised XLS files have:    date (M/YYYY) | nord | centre_et_sud | sahel
    """
    import openpyxl

    # --- Regional XLSX files (brebis, genisses, vaches) ---------------------
    regional_rows = [
        [2020, "janvier", 800.0, 820.0, 790.0],
        [2020, "février", 810.0, 830.0, 800.0],
        [2020, "mars", 820.0, 840.0, 810.0],
    ]
    for fname in _REGIONAL_FILES:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["annee", "mois", "nord", "sahel", "centre-et-sud"])
        for row in regional_rows:
            ws.append(row)
        wb.save(tmp_path / fname)

    # --- Meat XLSX -----------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "annee", "mois",
        "taurillon-maigre", "taurillon-engraisse", "agneau", "antenais", "caprin",
    ])
    for month, mois in enumerate(["janvier", "février", "mars"], start=0):
        ws.append([2020, mois, 12.0 + month, 14.0 + month, 18.0 + month,
                   16.0 + month, 10.0 + month])
    wb.save(tmp_path / _MEAT_FILE)

    # --- HTML-disguised XLS files (bovins, vaches gestantes) ----------------
    html = (
        "<html><body><table>\n"
        "<tr><th>date</th><th>nord</th><th>centre_et_sud</th><th>sahel</th></tr>\n"
        "<tr><td>1/2020</td><td>900.0</td><td>880.0</td><td>890.0</td></tr>\n"
        "<tr><td>2/2020</td><td>910.0</td><td>890.0</td><td>900.0</td></tr>\n"
        "<tr><td>3/2020</td><td>920.0</td><td>900.0</td><td>910.0</td></tr>\n"
        "</table></body></html>"
    )
    for fname in _XLS_HTML_FILES:
        (tmp_path / fname).write_text(html, encoding="utf-8")

    # --- Arabic fodder Excel files (tbn, qrt) --------------------------------
    _FODDER_FILES = [
        "تطور-أسعار-التبن.xlsx",
        "تطور-أسعار-القرط.xlsx",
    ]
    _FODDER_REGIONS = ["الشمال", "الساحل", "الوسط /ج"]
    for fname in _FODDER_FILES:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feuil1"
        for year in [2020, 2021, 2022]:
            ws.append([f"تطور أسعار لسنة {year}"] + [None] * 12)
            for region_ar in _FODDER_REGIONS:
                row = [region_ar] + ["10,0-8,0"] * 12
                ws.append(row)
        wb.save(tmp_path / fname)

    return tmp_path


# ---------------------------------------------------------------------------
# DB session mock
# ---------------------------------------------------------------------------


def _make_mock_session() -> AsyncMock:
    """Build an AsyncMock that satisfies all session call patterns in the repo."""
    session = AsyncMock()

    # execute() → object with .scalars().all() → [] and .scalar_one_or_none() → None
    result_mock = MagicMock()
    result_mock.scalars.return_value.all.return_value = []
    result_mock.scalar_one_or_none.return_value = None
    session.execute.return_value = result_mock

    # add() is sync in SQLAlchemy
    session.add = MagicMock()

    return session


@pytest.fixture
def mock_db_session() -> AsyncMock:
    return _make_mock_session()


# ---------------------------------------------------------------------------
# Async HTTP client with injected mock session
# ---------------------------------------------------------------------------


@pytest.fixture
async def async_client(mock_db_session: AsyncMock) -> AsyncClient:
    """httpx.AsyncClient bound to the FastAPI app with DB dependency overridden."""

    async def _override_session():
        yield mock_db_session

    app.dependency_overrides[get_async_session] = _override_session
    token = _make_test_token()
    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://test",
        headers={"Authorization": f"Bearer {token}"},
    ) as client:
        yield client
    app.dependency_overrides.clear()
